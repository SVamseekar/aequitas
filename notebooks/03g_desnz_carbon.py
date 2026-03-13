# %% [markdown]
# # 03g — DESNZ GHG Conversion Factors 2025
#
# **Purpose:** Extract transport CO2 emission factors from the DESNZ 2025
# condensed-set spreadsheet and save as a structured JSON for downstream use
# in 04e (economic appraisal / modal shift carbon savings).
#
# **Input:** `data/raw/carbon/ghg-conversion-factors-2025-condensed-set.xlsx`
# (1.8 MB, published June 2025)
#
# **Output:** `data/audit/desnz_carbon_factors.json`
#
# **Resolves:** CO2-001 (Bus CO2 per pax-km) and CO2-002 (Car CO2 per km)
# from ⚠️ Stale → ✅ Confirmed.

# %%
import json
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger

# %%
ROOT = Path("/Users/souravamseekarmarti/Projects/aequitas")
DATA = ROOT / "data"
DESNZ_PATH = DATA / "raw/carbon/ghg-conversion-factors-2025-condensed-set.xlsx"
OUT_PATH = DATA / "audit/desnz_carbon_factors.json"

assert DESNZ_PATH.exists(), f"DESNZ file not found: {DESNZ_PATH}"
logger.info(f"Loading workbook: {DESNZ_PATH.name} ({DESNZ_PATH.stat().st_size / 1e6:.1f} MB)")

# %% [markdown]
# ## 1. Load workbook — list ALL sheet names

# %%
wb = openpyxl.load_workbook(DESNZ_PATH, data_only=True)
logger.info(f"Workbook loaded: {len(wb.sheetnames)} sheets")

for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"  [{i:2d}] {name:40s}  rows={ws.max_row:>4d}  cols={ws.max_column:>3d}")

# %% [markdown]
# ## 2. Navigate to transport factors
#
# The relevant sheet for business travel (bus, rail, taxi) is **"Business travel- land"**
# (Scope 3). Car factors for company-owned vehicles are in **"Passenger vehicles"**
# (Scope 1), but the same "average car" value appears in "Business travel- land"
# for employee-owned vehicles.
#
# We use "Business travel- land" as our primary source because it contains both
# bus and car factors on the same sheet with consistent units, and the car values
# match the Passenger vehicles sheet exactly.

# %%
SHEET_NAME = "Business travel- land"
ws = wb[SHEET_NAME]
logger.info(f"Using sheet: '{SHEET_NAME}' ({ws.max_row} rows × {ws.max_column} cols)")

# Load into DataFrame for easier navigation
rows_raw = [list(r) for r in ws.iter_rows(values_only=True)]
df = pd.DataFrame(rows_raw)

# Confirm scope
scope_row = df.iloc[5]
scope_vals = [c for c in scope_row if c is not None and str(c).strip()]
logger.info(f"Scope info: {scope_vals}")

# %% [markdown]
# ## 3. Extract bus emission factors
#
# Bus factors are at rows 79-82 (1-indexed Excel rows), column D = total kg CO2e.
# Unit: passenger.km (i.e. per passenger-km, well-to-wheel including vehicle
# manufacture amortisation in Scope 3).

# %%
# --- Bus factors ---
# Cell references: D79 (local not London), D80 (London), D81 (average local), D82 (coach)
bus_factors = {}

bus_cells = {
    "local_bus_not_london": {"row": 79, "col": "D", "expected_type": "Local bus (not London)"},
    "local_london_bus": {"row": 80, "col": "D", "expected_type": "Local London bus"},
    "average_local_bus": {"row": 81, "col": "D", "expected_type": "Average local bus"},
    "coach": {"row": 82, "col": "D", "expected_type": "Coach"},
}

for key, info in bus_cells.items():
    type_cell = ws[f"B{info['row']}"].value
    unit_cell = ws[f"C{info['row']}"].value
    value_cell = ws[f"D{info['row']}"].value

    assert type_cell == info["expected_type"], (
        f"Cell B{info['row']} expected '{info['expected_type']}', got '{type_cell}'"
    )
    assert unit_cell == "passenger.km", f"Unit mismatch at C{info['row']}: {unit_cell}"
    assert isinstance(value_cell, (int, float)), f"Non-numeric value at D{info['row']}: {value_cell}"

    bus_factors[key] = {
        "value_kg_co2e_per_pax_km": round(float(value_cell), 5),
        "cell_ref": f"'{SHEET_NAME}'!D{info['row']}",
        "unit": "kg CO2e / passenger-km",
        "scope": "Scope 3 (well-to-wheel, includes vehicle manufacture amortisation)",
    }
    logger.info(f"Bus | {info['expected_type']:25s} | {value_cell:.5f} kg CO2e/pax-km | D{info['row']}")

print()
print("Bus emission factors extracted:")
for k, v in bus_factors.items():
    print(f"  {k}: {v['value_kg_co2e_per_pax_km']} kg CO2e/pax-km")

# %% [markdown]
# ## 4. Extract car emission factors
#
# Car factors in "Business travel- land" rows 47-53 (by size), column D = diesel
# total kg CO2e per vehicle-km. Row 53 = "Average car" diesel.
#
# Note: column D is the default (diesel) value. The sheet also has petrol (H),
# hybrid (L), etc. We extract diesel as the standard benchmark and also note
# the petrol and unknown-fuel values.

# %%
# --- Car factors ---
# Cell references: D47 (small), D49 (medium), D51 (large), D53 (average)
car_cells = {
    "small_car_diesel": {"row": 47, "col": "D", "expected_type": "Small car"},
    "medium_car_diesel": {"row": 49, "col": "D", "expected_type": "Medium car"},
    "large_car_diesel": {"row": 51, "col": "D", "expected_type": "Large car"},
    "average_car_diesel": {"row": 53, "col": "D", "expected_type": "Average car"},
}

car_factors = {}
for key, info in car_cells.items():
    type_cell = ws[f"B{info['row']}"].value
    unit_cell = ws[f"C{info['row']}"].value
    value_cell = ws[f"D{info['row']}"].value

    assert type_cell == info["expected_type"], (
        f"Cell B{info['row']} expected '{info['expected_type']}', got '{type_cell}'"
    )
    assert unit_cell == "km", f"Unit mismatch at C{info['row']}: {unit_cell}"
    assert isinstance(value_cell, (int, float)), f"Non-numeric value at D{info['row']}: {value_cell}"

    car_factors[key] = {
        "value_kg_co2e_per_km": round(float(value_cell), 5),
        "cell_ref": f"'{SHEET_NAME}'!D{info['row']}",
        "unit": "kg CO2e / vehicle-km",
        "fuel_type": "diesel",
        "scope": "Scope 3 (well-to-wheel)",
    }
    logger.info(f"Car | {info['expected_type']:15s} (diesel) | {value_cell:.5f} kg CO2e/km | D{info['row']}")

# Also extract petrol and unknown-fuel for average car
avg_car_petrol = ws["H53"].value
avg_car_unknown = ws["X53"].value

# Verify column headers for out-of-loop extractions (H=petrol, X=unknown fuel)
# Row 45 contains fuel-type labels: D='Diesel', H='Petrol', X='Unknown'
petrol_header = ws.cell(row=45, column=8).value  # column H
unknown_header = ws.cell(row=45, column=24).value  # column X
assert petrol_header is not None and "petrol" in str(petrol_header).lower(), (
    f"Column H header at row 45 unexpected: {petrol_header!r} — expected 'Petrol'"
)
assert unknown_header is not None and "unknown" in str(unknown_header).lower(), (
    f"Column X header at row 45 unexpected: {unknown_header!r} — expected 'Unknown'"
)
logger.info(f"Column header check: H45={petrol_header!r}, X45={unknown_header!r}")

car_factors["average_car_petrol"] = {
    "value_kg_co2e_per_km": round(float(avg_car_petrol), 5),
    "cell_ref": f"'{SHEET_NAME}'!H53",
    "unit": "kg CO2e / vehicle-km",
    "fuel_type": "petrol",
    "scope": "Scope 3 (well-to-wheel)",
}

car_factors["average_car_unknown_fuel"] = {
    "value_kg_co2e_per_km": round(float(avg_car_unknown), 5),
    "cell_ref": f"'{SHEET_NAME}'!X53",
    "unit": "kg CO2e / vehicle-km",
    "fuel_type": "unknown (weighted average)",
    "scope": "Scope 3 (well-to-wheel)",
}

logger.info(f"Car | Average car (petrol)  | {avg_car_petrol:.5f} kg CO2e/km | H53")
logger.info(f"Car | Average car (unknown) | {avg_car_unknown:.5f} kg CO2e/km | X53")

print()
print("Car emission factors extracted:")
for k, v in car_factors.items():
    print(f"  {k}: {v['value_kg_co2e_per_km']} kg CO2e/vehicle-km ({v['fuel_type']})")

# %% [markdown]
# ## 5. Extract rail emission factors (for modal comparison)

# %%
# --- Rail factors ---
# Cell references: D87 (national rail), D88 (international), D89 (light rail/tram), D90 (underground)
rail_cells = {
    "national_rail": {"row": 87, "col": "D", "expected_type": "National rail"},
    "international_rail": {"row": 88, "col": "D", "expected_type": "International rail"},
    "light_rail_and_tram": {"row": 89, "col": "D", "expected_type": "Light rail and tram"},
    "london_underground": {"row": 90, "col": "D", "expected_type": "London Underground"},
}

rail_factors = {}
for key, info in rail_cells.items():
    type_cell = ws[f"B{info['row']}"].value
    unit_cell = ws[f"C{info['row']}"].value
    value_cell = ws[f"D{info['row']}"].value

    assert type_cell == info["expected_type"], (
        f"Cell B{info['row']} expected '{info['expected_type']}', got '{type_cell}'"
    )
    assert unit_cell == "passenger.km", f"Unit mismatch at C{info['row']}: {unit_cell}"
    assert isinstance(value_cell, (int, float)), f"Non-numeric value at D{info['row']}: {value_cell}"

    rail_factors[key] = {
        "value_kg_co2e_per_pax_km": round(float(value_cell), 5),
        "cell_ref": f"'{SHEET_NAME}'!D{info['row']}",
        "unit": "kg CO2e / passenger-km",
        "scope": "Scope 3 (well-to-wheel)",
    }
    logger.info(f"Rail | {info['expected_type']:25s} | {value_cell:.5f} kg CO2e/pax-km | D{info['row']}")

print()
print("Rail emission factors extracted:")
for k, v in rail_factors.items():
    print(f"  {k}: {v['value_kg_co2e_per_pax_km']} kg CO2e/pax-km")

# %% [markdown]
# ## 6. Extract taxi factors (for completeness)

# %%
# --- Taxi factors ---
taxi_cells = {
    "regular_taxi": {"row": 71, "col": "D", "expected_type": "Regular taxi"},
    "black_cab": {"row": 73, "col": "D", "expected_type": "Black cab"},
}

taxi_factors = {}
for key, info in taxi_cells.items():
    type_cell = ws[f"B{info['row']}"].value
    unit_cell = ws[f"C{info['row']}"].value
    value_cell = ws[f"D{info['row']}"].value

    assert type_cell == info["expected_type"], (
        f"Cell B{info['row']} expected '{info['expected_type']}', got '{type_cell}'"
    )
    assert unit_cell == "passenger.km", f"Unit mismatch at C{info['row']}: {unit_cell}"

    taxi_factors[key] = {
        "value_kg_co2e_per_pax_km": round(float(value_cell), 5),
        "cell_ref": f"'{SHEET_NAME}'!D{info['row']}",
        "unit": "kg CO2e / passenger-km",
        "scope": "Scope 3 (well-to-wheel)",
    }
    logger.info(f"Taxi | {info['expected_type']:15s} | {value_cell:.5f} kg CO2e/pax-km | D{info['row']}")

# %% [markdown]
# ## 7. Car per passenger-km derivation
#
# DESNZ car factors are per **vehicle-km**. To compare with bus/rail (per pax-km),
# we divide by average occupancy. DfT National Travel Survey 2023 average car
# occupancy = **1.55 persons**.

# %%
DFT_CAR_OCCUPANCY = 1.55  # DfT NTS 2023

avg_car_diesel_per_km = car_factors["average_car_diesel"]["value_kg_co2e_per_km"]
avg_car_per_pax_km = round(avg_car_diesel_per_km / DFT_CAR_OCCUPANCY, 5)

logger.info(
    f"Car per pax-km (derived): {avg_car_diesel_per_km:.5f} / {DFT_CAR_OCCUPANCY} = "
    f"{avg_car_per_pax_km:.5f} kg CO2e/pax-km"
)

print(f"\nDerived: average car per passenger-km = {avg_car_per_pax_km} kg CO2e/pax-km")
print(f"  (= {avg_car_diesel_per_km} / {DFT_CAR_OCCUPANCY} occupancy)")

# %% [markdown]
# ## 8. Comparison with stale values
#
# The previous estimates in figures-registry were:
# - CO2-001: Bus = 0.0965 kg CO2e/pax-km (stale, likely from an older DEFRA year)
# - CO2-002: Car = 0.171 kg CO2e/km (stale)

# %%
STALE_BUS = 0.0965
STALE_CAR = 0.171

confirmed_bus = bus_factors["average_local_bus"]["value_kg_co2e_per_pax_km"]
confirmed_car = car_factors["average_car_diesel"]["value_kg_co2e_per_km"]

bus_delta_pct = ((confirmed_bus - STALE_BUS) / STALE_BUS) * 100
car_delta_pct = ((confirmed_car - STALE_CAR) / STALE_CAR) * 100

print("Comparison with stale values:")
print(f"  CO2-001 Bus: stale={STALE_BUS} → confirmed={confirmed_bus} "
      f"(delta={bus_delta_pct:+.1f}%)")
print(f"  CO2-002 Car: stale={STALE_CAR} → confirmed={confirmed_car} "
      f"(delta={car_delta_pct:+.1f}%)")

logger.info(f"CO2-001 bus delta: {bus_delta_pct:+.1f}% vs stale")
logger.info(f"CO2-002 car delta: {car_delta_pct:+.1f}% vs stale")

# %% [markdown]
# ## 9. Modal comparison table

# %%
modal_comparison = {
    "Walk / cycle": 0.0,
    "National rail": rail_factors["national_rail"]["value_kg_co2e_per_pax_km"],
    "Light rail / tram": rail_factors["light_rail_and_tram"]["value_kg_co2e_per_pax_km"],
    "London Underground": rail_factors["london_underground"]["value_kg_co2e_per_pax_km"],
    "Coach": bus_factors["coach"]["value_kg_co2e_per_pax_km"],
    "Average local bus": bus_factors["average_local_bus"]["value_kg_co2e_per_pax_km"],
    "Local bus (not London)": bus_factors["local_bus_not_london"]["value_kg_co2e_per_pax_km"],
    "Average car (per pax-km)": avg_car_per_pax_km,
    "Regular taxi": taxi_factors["regular_taxi"]["value_kg_co2e_per_pax_km"],
    "Black cab": taxi_factors["black_cab"]["value_kg_co2e_per_pax_km"],
}

print("\nModal CO2 comparison (kg CO2e per passenger-km, ascending):")
print("-" * 55)
for mode, val in sorted(modal_comparison.items(), key=lambda x: x[1]):
    bar = "#" * int(val * 200)
    print(f"  {mode:30s} {val:.5f}  {bar}")

# Car-to-bus modal shift saving
modal_shift_saving = avg_car_per_pax_km - confirmed_bus
print(f"\nModal shift saving (car → bus): {modal_shift_saving:.5f} kg CO2e/pax-km")
print(f"  = {modal_shift_saving * 1000:.2f} g CO2e per passenger-km switched")

# %% [markdown]
# ## 10. Build and save output JSON

# %%
output = {
    "source": "DESNZ UK Government GHG Conversion Factors for Company Reporting 2025",
    "file": "ghg-conversion-factors-2025-condensed-set.xlsx",
    "sheet": SHEET_NAME,
    "published": "June 2025",
    "factor_set": "Condensed set",
    "version": 1,
    "scope": "Scope 3 (well-to-wheel, includes upstream fuel production and vehicle manufacture amortisation)",
    "extraction_date": "2026-03-13",
    "extracted_by": "03g_desnz_carbon.ipynb",
    "bus": {
        "average_local_bus": {
            "value": confirmed_bus,
            "unit": "kg CO2e / passenger-km",
            "cell_ref": bus_factors["average_local_bus"]["cell_ref"],
        },
        "local_bus_not_london": {
            "value": bus_factors["local_bus_not_london"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": bus_factors["local_bus_not_london"]["cell_ref"],
        },
        "local_london_bus": {
            "value": bus_factors["local_london_bus"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": bus_factors["local_london_bus"]["cell_ref"],
        },
        "coach": {
            "value": bus_factors["coach"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": bus_factors["coach"]["cell_ref"],
        },
    },
    "car": {
        "average_car_diesel_per_vehicle_km": {
            "value": confirmed_car,
            "unit": "kg CO2e / vehicle-km",
            "cell_ref": car_factors["average_car_diesel"]["cell_ref"],
            "fuel_type": "diesel",
        },
        "average_car_petrol_per_vehicle_km": {
            "value": car_factors["average_car_petrol"]["value_kg_co2e_per_km"],
            "unit": "kg CO2e / vehicle-km",
            "cell_ref": car_factors["average_car_petrol"]["cell_ref"],
            "fuel_type": "petrol",
        },
        "average_car_unknown_fuel_per_vehicle_km": {
            "value": car_factors["average_car_unknown_fuel"]["value_kg_co2e_per_km"],
            "unit": "kg CO2e / vehicle-km",
            "cell_ref": car_factors["average_car_unknown_fuel"]["cell_ref"],
            "fuel_type": "unknown (weighted average across fleet)",
        },
        "average_car_per_passenger_km": {
            "value": avg_car_per_pax_km,
            "unit": "kg CO2e / passenger-km",
            "derivation": f"average_car_diesel ({confirmed_car}) / DfT NTS 2023 occupancy ({DFT_CAR_OCCUPANCY})",
        },
        "occupancy_assumption": {
            "value": DFT_CAR_OCCUPANCY,
            "source": "DfT National Travel Survey 2023",
        },
    },
    "rail": {
        "national_rail": {
            "value": rail_factors["national_rail"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": rail_factors["national_rail"]["cell_ref"],
        },
        "light_rail_and_tram": {
            "value": rail_factors["light_rail_and_tram"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": rail_factors["light_rail_and_tram"]["cell_ref"],
        },
        "london_underground": {
            "value": rail_factors["london_underground"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": rail_factors["london_underground"]["cell_ref"],
        },
    },
    "taxi": {
        "regular_taxi": {
            "value": taxi_factors["regular_taxi"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": taxi_factors["regular_taxi"]["cell_ref"],
        },
        "black_cab": {
            "value": taxi_factors["black_cab"]["value_kg_co2e_per_pax_km"],
            "unit": "kg CO2e / passenger-km",
            "cell_ref": taxi_factors["black_cab"]["cell_ref"],
        },
    },
    "baseline": {
        "walk": {"value": 0.0, "unit": "kg CO2e / passenger-km"},
        "cycle": {"value": 0.0, "unit": "kg CO2e / passenger-km"},
    },
    "modal_shift_car_to_bus": {
        "saving_per_pax_km": round(modal_shift_saving, 5),
        "unit": "kg CO2e / passenger-km",
        "car_basis": f"average_car_diesel_per_vehicle_km ({confirmed_car}) / {DFT_CAR_OCCUPANCY} occupancy",
        "bus_basis": "average_local_bus",
        "note": "Positive = CO2 saved by switching from car to bus",
    },
    "notes": [
        "Bus, rail, taxi, coach: kg CO2e per passenger-km (occupancy already embedded in factor)",
        "Car: kg CO2e per vehicle-km (divide by occupancy for per-passenger-km comparison)",
        "All factors are Scope 3 well-to-wheel (includes upstream fuel production)",
        "Modal shift CO2 saving = (car_per_pax_km - bus_per_pax_km) x distance_km x passengers",
        "DESNZ 2025 uses 2023 fleet composition data",
    ],
}

OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
with open(OUT_PATH, "w") as f:
    json.dump(output, f, indent=2)

logger.info(f"Saved: {OUT_PATH} ({OUT_PATH.stat().st_size:,} bytes)")
print(f"\nOutput saved: {OUT_PATH}")

# %% [markdown]
# ## 11. Validation summary

# %%
checks_passed = 0
checks_warned = 0
checks_failed = 0

def check(name: str, condition: bool, msg: str = "") -> None:
    """Run a validation check and print PASS/WARN/FAIL."""
    global checks_passed, checks_warned, checks_failed
    if condition:
        checks_passed += 1
        print(f"  PASS  {name}")
    else:
        checks_failed += 1
        print(f"  FAIL  {name}: {msg}")

def warn_check(name: str, condition: bool, msg: str = "") -> None:
    """Run a validation check — WARN instead of FAIL."""
    global checks_passed, checks_warned
    if condition:
        checks_passed += 1
        print(f"  PASS  {name}")
    else:
        checks_warned += 1
        print(f"  WARN  {name}: {msg}")


print("=" * 60)
print("VALIDATION SUMMARY — 03g DESNZ Carbon Factors")
print("=" * 60)

# Core extraction checks
check("CO2-001 bus avg extracted", confirmed_bus == 0.10385, f"got {confirmed_bus}")
check("CO2-002 car avg extracted", confirmed_car == 0.17304, f"got {confirmed_car}")
check("CO2-003 rail extracted",
      rail_factors["national_rail"]["value_kg_co2e_per_pax_km"] == 0.03546)

# Sanity range checks
check("Bus factor in range [0.05, 0.20]", 0.05 <= confirmed_bus <= 0.20)
check("Car factor in range [0.10, 0.30]", 0.10 <= confirmed_car <= 0.30)
check("Rail < bus (expected)", rail_factors["national_rail"]["value_kg_co2e_per_pax_km"] < confirmed_bus)
check("Bus < car per pax-km (expected)", confirmed_bus < avg_car_per_pax_km)
check("Coach < bus (expected)", bus_factors["coach"]["value_kg_co2e_per_pax_km"] < confirmed_bus)

# Stale value comparison — within 20% is acceptable
warn_check("Bus vs stale within 20%", abs(bus_delta_pct) < 20,
           f"delta={bus_delta_pct:+.1f}%")
warn_check("Car vs stale within 20%", abs(car_delta_pct) < 20,
           f"delta={car_delta_pct:+.1f}%")

# Output file checks
check("Output JSON exists", OUT_PATH.exists())
with open(OUT_PATH) as f:
    loaded = json.load(f)
check("JSON has bus section", "bus" in loaded)
check("JSON has car section", "car" in loaded)
check("JSON has rail section", "rail" in loaded)
check("JSON has source metadata", loaded.get("source", "").startswith("DESNZ"))

print()
print(f"Results: {checks_passed} PASS, {checks_warned} WARN, {checks_failed} FAIL")
if checks_failed == 0:
    logger.success("All validation checks passed")
else:
    logger.error(f"{checks_failed} checks FAILED")

# %% [markdown]
# ## 12. Key findings
#
# | ID | Metric | DESNZ 2025 Value | Stale Value | Delta | Cell Ref |
# |-----|--------|-----------------|-------------|-------|----------|
# | CO2-001 | Bus avg local (pax-km) | 0.10385 kg CO2e | 0.0965 | +7.6% | D81 |
# | CO2-002 | Car avg diesel (veh-km) | 0.17304 kg CO2e | 0.171 | +1.2% | D53 |
# | CO2-003 | National rail (pax-km) | 0.03546 kg CO2e | — | — | D87 |
#
# **Scope:** All factors are Scope 3 well-to-wheel (includes upstream fuel
# production and vehicle manufacture amortisation).
#
# **Modal shift saving:** Switching one passenger-km from car to bus saves
# approximately 0.78 g CO2e (car per pax-km 0.11164 minus bus 0.10385).
# The small gap reflects that average local bus (outside London) has relatively
# high per-passenger emissions due to low occupancy on many routes — a finding
# that reinforces the importance of frequency improvements in the policy model.
#
# **London vs non-London:** London buses emit 0.06875 vs 0.12525 outside London,
# reflecting higher occupancy in London. The average (0.10385) is occupancy-weighted.

# %%
print("\n03g_desnz_carbon COMPLETE")
